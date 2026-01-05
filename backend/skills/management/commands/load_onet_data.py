"""
Management command to load O*NET occupation and skills data.
"""
import os
from django.core.management.base import BaseCommand
from django.conf import settings
import openpyxl

from skills.models import Occupation, Skill, OccupationSkill


# Office-based SOC major groups
OFFICE_BASED_GROUPS = {
    '11',  # Management Occupations
    '13',  # Business and Financial Operations
    '15',  # Computer and Mathematical
    '23',  # Legal Occupations
    '27',  # Arts, Design, Entertainment, Sports, and Media
    '41',  # Sales and Related
    '43',  # Office and Administrative Support
}

# O*NET Skills to load (element IDs) - category must match model choices: knowledge, skill, ability, tool
SKILL_ELEMENTS = {
    '2.A.1.a': ('Reading Comprehension', 'Understanding written sentences and paragraphs in work-related documents.', 'skill'),
    '2.A.1.b': ('Active Listening', 'Giving full attention to what other people are saying, taking time to understand the points being made, asking questions as appropriate, and not interrupting at inappropriate times.', 'skill'),
    '2.A.1.c': ('Writing', 'Communicating effectively in writing as appropriate for the needs of the audience.', 'skill'),
    '2.A.1.d': ('Speaking', 'Talking to others to convey information effectively.', 'skill'),
    '2.A.1.e': ('Mathematics', 'Using mathematics to solve problems.', 'skill'),
    '2.A.1.f': ('Science', 'Using scientific rules and methods to solve problems.', 'skill'),
    '2.A.2.a': ('Critical Thinking', 'Using logic and reasoning to identify the strengths and weaknesses of alternative solutions, conclusions, or approaches to problems.', 'skill'),
    '2.A.2.b': ('Active Learning', 'Understanding the implications of new information for both current and future problem-solving and decision-making.', 'skill'),
    '2.A.2.c': ('Learning Strategies', 'Selecting and using training/instructional methods and procedures appropriate for the situation when learning or teaching new things.', 'skill'),
    '2.A.2.d': ('Monitoring', 'Monitoring/Assessing performance of yourself, other individuals, or organizations to make improvements or take corrective action.', 'skill'),
    '2.B.1.a': ('Social Perceptiveness', 'Being aware of others\' reactions and understanding why they react as they do.', 'skill'),
    '2.B.1.b': ('Coordination', 'Adjusting actions in relation to others\' actions.', 'skill'),
    '2.B.1.c': ('Persuasion', 'Persuading others to change their minds or behavior.', 'skill'),
    '2.B.1.d': ('Negotiation', 'Bringing others together and trying to reconcile differences.', 'skill'),
    '2.B.1.e': ('Instructing', 'Teaching others how to do something.', 'skill'),
    '2.B.1.f': ('Service Orientation', 'Actively looking for ways to help people.', 'skill'),
    '2.B.2.i': ('Complex Problem Solving', 'Identifying complex problems and reviewing related information to develop and evaluate options and implement solutions.', 'skill'),
    '2.B.3.a': ('Operations Analysis', 'Analyzing needs and product requirements to create a design.', 'skill'),
    '2.B.3.b': ('Technology Design', 'Generating or adapting equipment and technology to serve user needs.', 'skill'),
    '2.B.3.c': ('Equipment Selection', 'Determining the kind of tools and equipment needed to do a job.', 'skill'),
    '2.B.3.d': ('Installation', 'Installing equipment, machines, wiring, or programs to meet specifications.', 'skill'),
    '2.B.3.e': ('Programming', 'Writing computer programs for various purposes.', 'skill'),
    '2.B.3.g': ('Operation Monitoring', 'Watching gauges, dials, or other indicators to make sure a machine is working properly.', 'skill'),
    '2.B.3.h': ('Operation and Control', 'Controlling operations of equipment or systems.', 'skill'),
    '2.B.3.j': ('Equipment Maintenance', 'Performing routine maintenance on equipment and determining when and what kind of maintenance is needed.', 'skill'),
    '2.B.3.k': ('Troubleshooting', 'Determining causes of operating errors and deciding what to do about it.', 'skill'),
    '2.B.3.l': ('Repairing', 'Repairing machines or systems using the needed tools.', 'skill'),
    '2.B.3.m': ('Quality Control Analysis', 'Conducting tests and inspections of products, services, or processes to evaluate quality or performance.', 'skill'),
    '2.B.4.e': ('Judgment and Decision Making', 'Considering the relative costs and benefits of potential actions to choose the most appropriate one.', 'skill'),
    '2.B.4.g': ('Systems Analysis', 'Determining how a system should work and how changes in conditions, operations, and the environment will affect outcomes.', 'skill'),
    '2.B.4.h': ('Systems Evaluation', 'Identifying measures or indicators of system performance and the actions needed to improve or correct performance, relative to the goals of the system.', 'skill'),
    '2.B.5.a': ('Time Management', 'Managing one\'s own time and the time of others.', 'skill'),
    '2.B.5.b': ('Management of Financial Resources', 'Determining how money will be spent to get the work done, and accounting for these expenditures.', 'skill'),
    '2.B.5.c': ('Management of Material Resources', 'Obtaining and seeing to the appropriate use of equipment, facilities, and materials needed to do certain work.', 'skill'),
    '2.B.5.d': ('Management of Personnel Resources', 'Motivating, developing, and directing people as they work, identifying the best people for the job.', 'skill'),
}


class Command(BaseCommand):
    help = 'Load O*NET occupation and skills data for office-based jobs'

    def add_arguments(self, parser):
        parser.add_argument(
            '--all',
            action='store_true',
            help='Load ALL occupations, not just office-based ones',
        )
        parser.add_argument(
            '--skills-only',
            action='store_true',
            help='Only load/update skills, not occupations',
        )

    def handle(self, *args, **options):
        data_dir = os.path.join(settings.BASE_DIR, 'data')

        # Check if data files exist
        occ_file = os.path.join(data_dir, 'onet_occupations.xlsx')
        skills_file = os.path.join(data_dir, 'onet_skills.xlsx')
        zones_file = os.path.join(data_dir, 'onet_job_zones.xlsx')

        if not os.path.exists(occ_file):
            self.stderr.write(self.style.ERROR(f'Occupation data file not found: {occ_file}'))
            return

        # Load skills first
        self.load_skills()

        if options['skills_only']:
            return

        # Load job zones
        job_zones = {}
        if os.path.exists(zones_file):
            job_zones = self.load_job_zones(zones_file)

        # Load occupations
        load_all = options['all']
        self.load_occupations(occ_file, job_zones, load_all)

        # Load occupation-skill relationships
        if os.path.exists(skills_file):
            self.load_occupation_skills(skills_file, load_all)

        self.stdout.write(self.style.SUCCESS('O*NET data loaded successfully!'))

    def load_skills(self):
        """Load or update skills from predefined list."""
        self.stdout.write('Loading skills...')

        created_count = 0
        updated_count = 0

        for element_id, (name, description, category) in SKILL_ELEMENTS.items():
            skill, created = Skill.objects.update_or_create(
                element_id=element_id,
                defaults={
                    'name': name,
                    'description': description,
                    'category': category,
                }
            )
            if created:
                created_count += 1
            else:
                updated_count += 1

        self.stdout.write(f'  Created {created_count} skills, updated {updated_count}')

    def load_job_zones(self, zones_file):
        """Load job zone data."""
        self.stdout.write('Loading job zones...')

        wb = openpyxl.load_workbook(zones_file)
        ws = wb.active

        # Headers: O*NET-SOC Code, Title, Job Zone, Date, Domain Source
        job_zones = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = row[0]
            job_zone = row[2]  # Job Zone is at index 2
            if code and job_zone:
                job_zones[code] = int(job_zone)

        self.stdout.write(f'  Loaded {len(job_zones)} job zone mappings')
        return job_zones

    def load_occupations(self, occ_file, job_zones, load_all=False):
        """Load occupations from O*NET data."""
        self.stdout.write('Loading occupations...')

        wb = openpyxl.load_workbook(occ_file)
        ws = wb.active

        created_count = 0
        updated_count = 0
        skipped_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            code, title, description = row[0], row[1], row[2]

            if not code:
                continue

            # Filter to office-based occupations unless --all is specified
            major_group = code[:2]
            if not load_all and major_group not in OFFICE_BASED_GROUPS:
                skipped_count += 1
                continue

            job_zone = job_zones.get(code, 3)  # Default to zone 3 if not found

            occ, created = Occupation.objects.update_or_create(
                onet_soc_code=code,
                defaults={
                    'title': title,
                    'description': description or '',
                    'job_zone': job_zone,
                }
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

        self.stdout.write(f'  Created {created_count} occupations, updated {updated_count}, skipped {skipped_count}')

    def load_occupation_skills(self, skills_file, load_all=False):
        """Load occupation-skill relationships."""
        self.stdout.write('Loading occupation-skill relationships...')

        wb = openpyxl.load_workbook(skills_file)
        ws = wb.active

        # Get column indices
        headers = [cell.value for cell in ws[1]]
        code_idx = headers.index('O*NET-SOC Code')
        element_idx = headers.index('Element ID')
        scale_idx = headers.index('Scale ID')
        value_idx = headers.index('Data Value')

        # Cache occupations and skills
        occupations = {o.onet_soc_code: o for o in Occupation.objects.all()}
        skills = {s.element_id: s for s in Skill.objects.all()}

        created_count = 0
        skipped_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            code = row[code_idx]
            element_id = row[element_idx]
            scale_id = row[scale_idx]
            value = row[value_idx]

            # Only process importance scale (IM)
            if scale_id != 'IM':
                continue

            # Skip if occupation or skill not in our data
            if code not in occupations:
                skipped_count += 1
                continue
            if element_id not in skills:
                continue

            occ = occupations[code]
            skill = skills[element_id]

            # Create or update the relationship
            occ_skill, created = OccupationSkill.objects.update_or_create(
                occupation=occ,
                skill=skill,
                defaults={
                    'importance': float(value) if value else 0,
                    'level': 0,  # Will be updated with LV scale data
                }
            )

            if created:
                created_count += 1

        # Now update with level data
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = row[code_idx]
            element_id = row[element_idx]
            scale_id = row[scale_idx]
            value = row[value_idx]

            if scale_id != 'LV':
                continue

            if code not in occupations or element_id not in skills:
                continue

            try:
                occ_skill = OccupationSkill.objects.get(
                    occupation=occupations[code],
                    skill=skills[element_id]
                )
                occ_skill.level = float(value) if value else 0
                occ_skill.save(update_fields=['level'])
            except OccupationSkill.DoesNotExist:
                pass

        self.stdout.write(f'  Created {created_count} occupation-skill links, skipped {skipped_count}')
